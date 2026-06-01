#!/usr/bin/env python3
# -*- coding: utf-8 -*-

# Usage: python 1_linkeranalysis.py linkers.sdf

import argparse
import os
import sys
import tempfile
import shutil
from collections import Counter, deque
import pandas as pd
import numpy as np
from rdkit import Chem
from rdkit.Chem import Draw, Descriptors, rdMolDescriptors
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import plotly.graph_objects as go
from plotly.subplots import make_subplots
from tqdm import tqdm


# SETTINGS
CATEGORY_COLORS = {
    "Aliphatic": "#F7C96A",
    "Ether": "#66B5D9",
    "Amide_Amine": "#66D1B0",
    "Functionalised": "#F7EE99",
    "Alicyclic_Heterocycle": "#C5D6FF",
    "Aromatic_Heteroaromatic": "#E88F55",
    "Triazole_Click": "#E3A9CA",
    "Rigid_Alkyne_Alkene": "#777777",
    "Other": "#CCCCCC",
}


# CLASSIFICATION CORE

def has_aromatic_beyond_triazole(mol):
    tri = ["n1ccnn1", "c1nn[nH]n1"]
    tri_atoms = set()
    for s in tri:
        p = Chem.MolFromSmarts(s)
        for m in mol.GetSubstructMatches(p):
            tri_atoms.update(m)
    for a in mol.GetAtoms():
        if a.GetIsAromatic() and a.GetIdx() not in tri_atoms:
            return True
    return False


def branch_depth(mol, start, block):
    visited = {block}
    q = deque([(start, 0)])
    md = 0
    while q:
        idx, d = q.popleft()
        a = mol.GetAtomWithIdx(idx)
        if a.GetAtomicNum() > 1:
            md = max(md, d + 1)
        for nb in a.GetNeighbors():
            j = nb.GetIdx()
            if j not in visited:
                visited.add(j)
                q.append((j, d + 1))
    return md


def max_depth_c(mol, center, block):
    ds = []
    for nb in center.GetNeighbors():
        if nb.GetIdx() != block and nb.GetAtomicNum() == 6:
            ds.append(branch_depth(mol, nb.GetIdx(), center.GetIdx()))
    return max(ds) if ds else 0


def longest_sp3_chain(mol):
    eligible = [
        a.GetIdx() for a in mol.GetAtoms()
        if a.GetAtomicNum() == 6
        and a.GetHybridization() == Chem.HybridizationType.SP3
        and not a.GetIsAromatic()
        and not a.IsInRing()
    ]
    if not eligible:
        return 0

    S = set(eligible)
    adj = {i: [] for i in eligible}

    for i in eligible:
        for nb in mol.GetAtomWithIdx(i).GetNeighbors():
            if nb.GetIdx() in S:
                adj[i].append(nb.GetIdx())

    def bfs(n):
        vis = {n}
        q = deque([(n, 0)])
        far, dist = n, 0
        while q:
            x, d = q.popleft()
            if d > dist:
                far, dist = x, d
            for w in adj[x]:
                if w not in vis:
                    vis.add(w)
                    q.append((w, d + 1))
        return far, dist

    visited, mlen = set(), 0
    for i in eligible:
        if i in visited:
            continue
        comp = []
        st = [i]
        while st:
            v = st.pop()
            if v not in comp:
                comp.append(v)
                for w in adj[v]:
                    st.append(w)
        visited |= set(comp)
        u, _ = bfs(comp[0])
        v, dist = bfs(u)
        mlen = max(mlen, dist + 1)
    return mlen


def classify_hits(mol, min_depth):
    hits = []

    sm = {
        "Functionalised": ["SS", "N=N", "[N]=[N]=[N]", "OC(=O)O", "CSC(=O)"],
        "Triazole_Click": ["n1ccnn1", "c1nn[nH]n1"],
        "Ether": ["OCCO", "COCO", "COCCO", "[C;!$(C=O)]-O-[C;!$(C=O)]"],
        "Alicyclic_Heterocycle": ["N1CCNCC1", "N1CCOCC1", "N1CCCCC1", "O1CCCC1", "N1CCCC1"],
        "Rigid_Alkyne_Alkene": ["C#C", "[C;!$(C~C=O)]=[C;!$(C~C=O)]"],
    }

    for cat, patts in sm.items():
        for s in patts:
            p = Chem.MolFromSmarts(s)
            if p and mol.HasSubstructMatch(p):
                hits.append(cat)
                break

    if has_aromatic_beyond_triazole(mol):
        hits.append("Aromatic_Heteroaromatic")

    if longest_sp3_chain(mol) >= 3:
        hits.append("Aliphatic")

    # internal amide-like
    internal = False
    for c in mol.GetAtoms():
        if c.GetAtomicNum() != 6:
            continue
        o_dbl = [b for b in c.GetBonds()
                 if b.GetBondTypeAsDouble() == 2 and b.GetOtherAtom(c).GetAtomicNum() == 8]
        n_sng = [b for b in c.GetBonds()
                 if b.GetBondTypeAsDouble() == 1 and b.GetOtherAtom(c).GetAtomicNum() == 7]
        if o_dbl and n_sng:
            n_atom = n_sng[0].GetOtherAtom(c)
            d1 = max_depth_c(mol, c, n_atom.GetIdx())
            d2 = max_depth_c(mol, n_atom, c.GetIdx())
            if d1 >= min_depth and d2 >= min_depth:
                internal = True
                break

    if internal:
        hits.append("Amide_Amine")

    out, seen = [], set()
    for h in hits:
        if h not in seen:
            out.append(h)
            seen.add(h)

    if not out:
        return "Other"
    if len(out) == 1:
        return out[0]
    return "Hybrid(" + " + ".join(out) + ")"


# PROPERTIES CALCULATION

def calc_props(mol):
    return {
        "heavy_atoms": sum(1 for a in mol.GetAtoms() if a.GetAtomicNum() > 1),
        "rot_bonds": rdMolDescriptors.CalcNumRotatableBonds(mol),
        "tpsa": rdMolDescriptors.CalcTPSA(mol),
        "clogp": Descriptors.MolLogP(mol),
        "fsp3": rdMolDescriptors.CalcFractionCSP3(mol),
        "rings": rdMolDescriptors.CalcNumRings(mol),
        "ar_rings": rdMolDescriptors.CalcNumAromaticRings(mol),
    }


# IMAGE EXPORT

def depict_to_png(mol, out_png):
    try:
        m2 = Chem.RemoveHs(mol)
        Draw.MolToImage(m2, size=(240, 180)).save(out_png)
        return True
    except:
        return False


# HTML SUMMARY REPORT

def generate_html_report(df, out_html):
    import plotly.graph_objects as go
    from plotly.subplots import make_subplots
    import numpy as np

    props = ["heavy_atoms", "rot_bonds", "tpsa", "clogp",
             "fsp3", "rings", "ar_rings"]

    xr, yr = {}, {}
    for p in props:
        col = df[p].dropna()
        xr[p] = (float(col.min()), float(col.max()))
        hist, _ = np.histogram(col, bins=30)
        yr[p] = (0, float(hist.max()) * 1.15)

    def split_categories(cat):
        if cat.startswith("Hybrid("):
            inside = cat[len("Hybrid("):-1]
            return [x.strip() for x in inside.split("+") if x.strip()]
        return [cat]

    cat_counts = Counter()
    for c in df["category"]:
        for x in split_categories(c):
            cat_counts[x] += 1

    cat_series = pd.Series(cat_counts).sort_values(ascending=False)

    fig = make_subplots(
        rows=2, cols=7,
        specs=[
            [{"type": "bar", "colspan": 7}, None, None, None, None, None, None],
            [{"type": "histogram"} for _ in range(7)]
        ],
        subplot_titles=["Category frequencies"] + props,
        vertical_spacing=0.1,
        horizontal_spacing=0.05,
    )

    # Barplot
    fig.add_trace(
        go.Bar(
            x=cat_series.index,
            y=cat_series.values,
            marker=dict(
                color=[CATEGORY_COLORS.get(c, "#888") for c in cat_series.index],
                line=dict(color="black", width=1)
            ),
            opacity=0.95,
        ),
        row=1, col=1
    )

    # Histograms
    for i, p in enumerate(props):
        fig.add_trace(
            go.Histogram(
                x=df[p],
                nbinsx=30,
                marker=dict(color="rgba(130,130,130,0.8)",
                            line=dict(width=1, color="black")),
                opacity=0.75,
            ),
            row=2, col=i + 1
        )
        fig.update_xaxes(range=xr[p], row=2, col=i + 1)
        fig.update_yaxes(range=yr[p], row=2, col=i + 1)

    # JS callback
    data_json = df.to_json(orient="records")
    js = f"""
var gd=document.getElementById('linker-report');
var allData={data_json};
var active=null;

function matches(cat,label){{
    if(cat.startsWith("Hybrid(")) return cat.indexOf(label)!==-1;
    return cat===label;
}}

function filterBy(label){{
    return allData.filter(d=>matches(d.category,label));
}}

gd.on('plotly_click',function(evt){{
    if(!evt || !evt.points.length) return;
    var label=evt.points[0].x;

    if(active===label){{
        active=null;
        Plotly.restyle(gd,{{
            x:[{",".join(["allData.map(d=>d."+p+")" for p in props])}]
        }},[{",".join([str(i) for i in range(1,8)])}]);
        Plotly.restyle(gd,{{opacity:[0.95]}},[0]);
        return;
    }}

    active=label;
    var f=filterBy(label);

    Plotly.restyle(gd,{{
        x:[{",".join(["f.map(d=>d."+p+")" for p in props])}]
    }},[{",".join([str(i) for i in range(1,8)])}]);

    var xs=gd.data[0].x;
    var op=[];
    for(var i=0;i<xs.length;i++) op.push(xs[i]===label?1.0:0.25);
    Plotly.restyle(gd,{{opacity:[op]}},[0]);
}});
"""

    fig.write_html(
        out_html,
        include_plotlyjs="cdn",
        full_html=True,
        div_id="linker-report",
        post_script=js
    )


# SPLIT CATEGORIES FOR COUNTS
def split_categories(cat):
    if cat.startswith("Hybrid("):
        inside = cat[len("Hybrid("):-1]
        return [x.strip() for x in inside.split("+") if x.strip()]
    return [cat]

# MAIN
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("sdf")
    ap.add_argument("--min-internal-depth", type=int, default=2)
    args = ap.parse_args()

    if not os.path.exists(args.sdf):
        sys.exit("SDF not found.")

    suppl = Chem.SDMolSupplier(args.sdf, removeHs=False)
    mols = [m for m in suppl if m]

    rows = []
    print(f"Processing {len(mols)} molecules...")
    for i, mol in enumerate(tqdm(mols, desc="Classifying")):
        name = mol.GetProp("_Name") if mol.HasProp("_Name") else f"mol_{i:05d}"
        smi = Chem.MolToSmiles(Chem.RemoveHs(mol), canonical=True)
        cat = classify_hits(mol, args.min_internal_depth)
        props = calc_props(mol)

        r = {"name": name, "smiles": smi, "category": cat}
        r.update(props)
        rows.append(r)

    df = pd.DataFrame(rows)

    # Save CSV + Excel
    df.to_csv("classified_linkers.csv", index=False)
    df.to_excel("classified_linkers.xlsx", index=False, sheet_name="linkers")

    wb = load_workbook("classified_linkers.xlsx")
    ws = wb["linkers"]

    ws.insert_cols(2)
    ws["B1"].value = "image"
    col_letter = get_column_letter(2)
    ws.column_dimensions[col_letter].width = 40

    tmp = tempfile.mkdtemp()
    try:
        print("Generating + inserting PNGs...")
        for r, mol in enumerate(tqdm(mols, desc="Images"), 2):
            png = os.path.join(tmp, f"{r - 1:05d}.png")
            if depict_to_png(mol, png):
                img = XLImage(png)
                img.width, img.height = 240, 180
                ws.add_image(img, f"{col_letter}{r}")
                ws.row_dimensions[r].height = 180 * 0.75

        wb.save("classified_linkers.xlsx")
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


    os.makedirs("reports", exist_ok=True)
    shutil.move("classified_linkers.csv", "reports/classified_linkers.csv")

    # 1) category_counts.csv
    cat_counts = Counter()
    for c in df["category"]:
        for x in split_categories(c):
            cat_counts[x] += 1

    pd.DataFrame(
        {"category": list(cat_counts.keys()),
         "count": list(cat_counts.values())}
    ).to_csv("reports/category_counts.csv", index=False)

    # 2) data_by_category.csv (expanded hybrids)
    expanded = []
    for _, row in df.iterrows():
        cats = split_categories(row["category"])
        for c in cats:
            newrow = row.copy()
            newrow["single_category"] = c
            expanded.append(newrow)

    df_expanded = pd.DataFrame(expanded)
    df_expanded.to_csv("reports/data_by_category.csv", index=False)

    # Generate HTML
    generate_html_report(df, "reports/summary.html")

    # 3) category_map.csv — True/False per pure category
    BASE_CATEGORIES = [
        "Aliphatic",
        "Ether",
        "Amide_Amine",
        "Functionalised",
        "Alicyclic_Heterocycle",
        "Aromatic_Heteroaromatic",
        "Triazole_Click",
        "Rigid_Alkyne_Alkene",
    ]

    map_rows = []

    for _, row in df.iterrows():
        cats = split_categories(row["category"])
        entry = {
            "name": row["name"],
            "smiles": row["smiles"],
            "category": row["category"],
        }

        for c in BASE_CATEGORIES:
            entry[c] = (c in cats)

        map_rows.append(entry)

    df_map = pd.DataFrame(map_rows)
    df_map.to_csv("reports/category_map.csv", index=False)
    
    print("Generated outputs in 'reports/' directory.")
    print("Generated classified_linkers.xlsx with images.")


if __name__ == "__main__":
    main()